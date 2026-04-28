"""
config/loader.py
================
Entry point for loading all WES model inputs from an .xlsx workbook.

Expected tab structure (one tab per section):

  SURVEILLANCE  — flat key/value table  (Parameter | Value | Unit | Notes)
  CONSTANTS     — flat key/value table
  OVERHEAD      — flat key/value table  (Parameter | Value (INR) | Frequency | Notes)
  ANNUALIZATION — flat key/value table  (Parameter | Value | Unit | Notes)

  HR            — wide table, one step per row
                  Step Key | Role | Num Personnel | Annual Salary (INR) |
                  Weekly WES Time (min) | Notes

  EQUIPMENT     — wide table, one item per row
                  Equipment Key | Step | Unit Cost (INR) | Avg Life (yrs) |
                  Num Units | Capacity Type |
                  Batch Size / Eq. Volume (L) | Time/Batch (min) or Sample Vol (L) |
                  Runs/Week | Shared?
                  (Columns G/H interpretation depends on Capacity Type)

  CONSUMABLES   — wide table, one item per row
                  Consumable Key | Step | Unit Cost (INR) | Qty/Sample |
                  Batch Size | Unit of Measure | Notes
                  (Qty/Sample "derived" → None; Batch Size "—" → None)

Row-skip rules (same for every tab):
  • First cell is blank                        → skip
  • First cell has a space AND does not start  → skip  (title / description /
    with "──"                                           column-header row)
  • First cell equals "Parameter" (any case)   → skip  (column-header row)
  • First cell starts with "──"                → section divider
      – flat tabs  : always skip (decorative only)
      – wide tables: skip if the extracted name has spaces (decorative group
                     header); otherwise treat as a regular data row (the
                     section name itself is NOT used as a key here)

Fuel note
─────────
consumable_calculator.py reads two fields from the "fuel" consumable entry that
are not columns in the CONSUMABLES tab:
  • total_distance_per_week_km
  • fuel_consumption_km_per_litre
Add both to your SURVEILLANCE tab.  The loader will copy them into the fuel
entry automatically.  Missing either key raises a clear ValueError.

Requires openpyxl:  pip install openpyxl
"""

import re

from utils.sheet_reader import is_url, resolve_input_sheet

_ALL_TABS = {
    "SURVEILLANCE", "CONSTANTS", "HR",
    "EQUIPMENT", "CONSUMABLES", "OVERHEAD", "ANNUALIZATION",
}


# ── Cell-value helpers ───────────────────────────────────────────────────────

def _coerce(val):
    """Convert an Excel cell value to the right Python scalar.

    Special strings:
      "—" / "–" / "-"         → None   (not-applicable marker)
      "derived" / "none" / "" → None   (dynamic or absent value)
      "True" / "False"        → bool
      otherwise try int, float, then keep as str
    """
    if val is None:
        return None
    # bool must be checked before int — bool is a subclass of int
    if isinstance(val, bool):
        return val
    if isinstance(val, (int, float)):
        return val
    s = str(val).strip()
    if not s or s.lower() in ("none", "derived", "n/a"):
        return None
    if s in ("—", "–", "-"):
        return None
    if s.lower() == "true":
        return True
    if s.lower() == "false":
        return False
    try:
        return int(s)
    except ValueError:
        pass
    try:
        return float(s)
    except ValueError:
        pass
    return s


def _num(val):
    """Coerce to a number, stripping trailing unit labels.

    Examples: "600 L" → 600.0,  "0.12 L" → 0.12,  "—" → None
    """
    if val is None:
        return None
    if isinstance(val, bool):
        return None
    if isinstance(val, (int, float)):
        return val
    s = str(val).strip()
    if not s or s in ("—", "–", "-", "derived", "n/a"):
        return None
    # Strip trailing unit label e.g. "600 L" → "600", "0.12 L" → "0.12"
    stripped = re.sub(r"\s+[A-Za-z/]+$", "", s).strip()
    try:
        return int(stripped)
    except ValueError:
        pass
    try:
        return float(stripped)
    except ValueError:
        pass
    return None


def _first(row) -> str:
    if not row:
        return ""
    return str(row[0].value).strip() if row[0].value is not None else ""


def _cell(row, i):
    return row[i].value if len(row) > i else None


def _is_skip(first: str) -> bool:
    """True for blank, title/description, and column-header rows."""
    if not first:
        return True
    # Multi-word text that is not a section divider → title or description line
    if " " in first and not first.startswith("──") and not first.startswith("--"):
        return True
    if first.lower() in ("parameter", "param"):
        return True
    return False


def _is_section_header(first: str) -> bool:
    return first.startswith("──") or first.startswith("--")


# ── Sheet parsers ────────────────────────────────────────────────────────────

def _parse_flat(ws) -> dict:
    """Flat key/value sheet: Parameter (col A) → Value (col B)."""
    result = {}
    for row in ws.iter_rows(min_col=1, max_col=4):
        f = _first(row)
        if _is_skip(f) or _is_section_header(f):
            continue
        result[f] = _coerce(_cell(row, 1))
    return result


def _parse_hr(ws) -> dict:
    """
    Wide-table HR sheet — one step per row:
      Step Key | Role | Num Personnel | Annual Salary (INR) |
      Weekly WES Time (min) | Notes
    """
    result = {}
    for row in ws.iter_rows(min_col=1, max_col=6):
        f = _first(row)
        if _is_skip(f) or _is_section_header(f):
            continue
        result[f] = {
            "role":                         _coerce(_cell(row, 1)),
            "num_personnel":                _coerce(_cell(row, 2)),
            "annual_salary_inr":            _coerce(_cell(row, 3)),
            "weekly_time_contribution_min": _coerce(_cell(row, 4)),
        }
    return result


def _parse_equipment(ws) -> dict:
    """
    Wide-table EQUIPMENT sheet — one item per row:
      Equipment Key | Step | Unit Cost (INR) | Avg Life (yrs) | Num Units |
      Capacity Type | col G | col H | Runs/Week | Shared?

    col G / col H meaning depends on Capacity Type:
      "time"      → batch_size               | time_per_batch_min
      "volume"    → equipment_volume_litres  | sample_volume_litres
      "misc_capex"→ (ignored)                | (ignored)

    "—" values in optional numeric columns → None (not applicable).
    "600 L" style values → numeric (unit label stripped).
    Section headers with spaces (e.g. "── Sample Collection & Transportation")
    are decorative and skipped.
    """
    result = {}
    for row in ws.iter_rows(min_col=1, max_col=10):
        f = _first(row)
        if _is_skip(f):
            continue
        if _is_section_header(f):
            continue  # decorative group headers (all have spaces after ──)

        cap_type = _coerce(_cell(row, 5))
        col6     = _num(_cell(row, 6))
        col7     = _num(_cell(row, 7))

        item = {
            "step":           _coerce(_cell(row, 1)),
            "unit_cost_inr":  _coerce(_cell(row, 2)),
            "avg_life_years": _coerce(_cell(row, 3)),
            "num_units":      _coerce(_cell(row, 4)),
            "capacity_type":  cap_type,
        }

        if cap_type == "time":
            item["batch_size"]         = col6
            item["time_per_batch_min"] = col7
            item["runs_per_week"]      = _coerce(_cell(row, 8))
        elif cap_type == "volume":
            item["equipment_volume_litres"] = col6
            item["sample_volume_litres"]    = col7
        # misc_capex: no extra fields required by the calculator

        shared = _coerce(_cell(row, 9))
        if shared is not None:
            item["shared_across_pathogens"] = shared

        result[f] = item
    return result


def _parse_consumables(ws) -> dict:
    """
    Wide-table CONSUMABLES sheet — one item per row.

    Standard columns (fixed mapping):
      Consumable Key | Step | Unit Cost (INR) | Qty/Sample |
      Batch Size | Unit of Measure | Notes

    Extra columns (col 7 onward, snake_case headers):
      Any column whose header is already snake_case is added to the item dict
      using the header as the key.  Blank / None values are omitted.
      Example: fuel_consumption_km_per_litre → stored only for rows that
      have a value; silently skipped for all others.

    Qty/Sample "derived" → None  (fuel, ice_pack — computed from other inputs)
    Batch Size "—"       → None  (not a batch consumable)
    Section headers with spaces are decorative and skipped.
    """
    # Standard columns: index → Python field name (None = skip)
    # Notes is intentionally absent — it is excluded by name in the header scan
    # below, so it can sit at any position (including last) without breaking the
    # extra-column detection.
    _STANDARD = {
        0: None,              # Consumable Key — used as dict key
        1: "step",
        2: "unit_cost_inr",
        3: "qty_per_sample",  # "derived" → None
        4: "batch_size",      # "—" → None
        5: "unit",
    }

    # Read all rows to find the header row first, then parse data rows
    all_rows = list(ws.iter_rows(min_col=1))

    # Find header row: first row whose first cell is "Consumable Key" (case-insensitive)
    header_row   = None
    extra_cols   = {}  # col_index → field_name for columns beyond the standard set
    for row in all_rows:
        f = _first(row)
        if f.lower() in ("consumable key", "consumablekey"):
            header_row = row
            for i, cell in enumerate(row):
                if i in _STANDARD:
                    continue
                h = str(cell.value).strip() if cell.value else ""
                # Only treat as an extra field if the header is snake_case
                if h and " " not in h and h.lower() not in ("notes", ""):
                    extra_cols[i] = h
            break

    result = {}
    for row in all_rows:
        f = _first(row)
        if _is_skip(f):
            continue
        if _is_section_header(f):
            continue
        if header_row is not None and row is header_row:
            continue

        item = {
            "step":           _coerce(_cell(row, 1)),
            "unit_cost_inr":  _coerce(_cell(row, 2)),
            "qty_per_sample": _coerce(_cell(row, 3)),
            "batch_size":     _coerce(_cell(row, 4)),
            "unit":           _coerce(_cell(row, 5)),
        }

        # Attach any extra column values (e.g. fuel_consumption_km_per_litre)
        for col_idx, field_name in extra_cols.items():
            val = _coerce(_cell(row, col_idx))
            if val is not None:
                item[field_name] = val

        result[f] = item
    return result


# ── Workbook loading ─────────────────────────────────────────────────────────

def _load_workbook(path: str):
    try:
        import openpyxl
    except ImportError:
        raise ImportError(
            "Reading Excel input sheets requires openpyxl.\n"
            "Install it with:  pip install openpyxl"
        )
    # read_only for speed; data_only to get computed values not raw formulas
    return openpyxl.load_workbook(path, read_only=True, data_only=True)


def _get_sheet(wb, name: str):
    """Case-insensitive tab lookup; returns None when the tab is absent."""
    upper = name.upper()
    for sheet_name in wb.sheetnames:
        if sheet_name.upper() == upper:
            return wb[sheet_name]
    return None


def _read_workbook(path: str) -> dict:
    wb = _load_workbook(path)

    missing = [tab for tab in _ALL_TABS if _get_sheet(wb, tab) is None]
    if missing:
        raise ValueError(
            f"Missing required tab(s) in workbook: {', '.join(sorted(missing))}\n"
            f"Expected tabs: {', '.join(sorted(_ALL_TABS))}"
        )

    result = {
        "SURVEILLANCE":  _parse_flat(_get_sheet(wb, "SURVEILLANCE")),
        "CONSTANTS":     _parse_flat(_get_sheet(wb, "CONSTANTS")),
        "HR":            _parse_hr(_get_sheet(wb, "HR")),
        "EQUIPMENT":     _parse_equipment(_get_sheet(wb, "EQUIPMENT")),
        "CONSUMABLES":   _parse_consumables(_get_sheet(wb, "CONSUMABLES")),
        "OVERHEAD":      _parse_flat(_get_sheet(wb, "OVERHEAD")),
        "ANNUALIZATION": _parse_flat(_get_sheet(wb, "ANNUALIZATION")),
    }

    # ── Fuel injection ───────────────────────────────────────────────────────
    # consumable_calculator._fuel_cost_per_sample() needs two extra fields
    # inside the fuel consumable dict:
    #
    #   total_distance_per_week_km   — derived automatically:
    #       EQUIPMENT["vehicle"]["time_per_batch_min"] × SURVEILLANCE["vehicle_speed_km_per_min"]
    #       e.g. 235 × 0.6667 ≈ 156.67 km  (no sheet change needed)
    #
    #   fuel_consumption_km_per_litre — read from the fuel row in the CONSUMABLES
    #       tab (extra column at the end of the row, e.g. value = 25)
    if "fuel" in result["CONSUMABLES"]:
        surv  = result["SURVEILLANCE"]
        equip = result["EQUIPMENT"]
        fuel  = result["CONSUMABLES"]["fuel"]

        # Derive total weekly driving distance from vehicle equipment entry
        vehicle       = equip.get("vehicle", {})
        vehicle_time  = vehicle.get("time_per_batch_min")
        vehicle_speed = surv.get("vehicle_speed_km_per_min")
        if vehicle_time is None or vehicle_speed is None:
            raise ValueError(
                "Cannot compute 'total_distance_per_week_km' for fuel cost. "
                "Ensure the EQUIPMENT tab has a 'vehicle' row with 'time_per_batch_min' "
                "and the SURVEILLANCE tab has 'vehicle_speed_km_per_min'."
            )
        fuel["total_distance_per_week_km"] = vehicle_time * vehicle_speed

        # fuel_consumption_km_per_litre must come from the CONSUMABLES sheet
        # as an extra column on the fuel row
        if "fuel_consumption_km_per_litre" not in fuel or fuel["fuel_consumption_km_per_litre"] is None:
            raise ValueError(
                "The fuel row in the CONSUMABLES tab is missing 'fuel_consumption_km_per_litre'. "
                "Add it as an extra column at the end of the CONSUMABLES sheet "
                "(e.g. column header = fuel_consumption_km_per_litre, value = 25)."
            )


    return result


# ── Public API ───────────────────────────────────────────────────────────────

def load_all_inputs(source: str) -> dict:
    """
    Resolve *source* (local path or Google Sheets URL), parse the workbook,
    and return all seven config dicts.

    Raises ValueError if any of the 7 required tabs is absent, or if the
    SURVEILLANCE tab is missing the fuel transport parameters.

    Extra keys:
        _resolved_path  local path that was actually used
        _is_temp        True when a temp file was downloaded and must be deleted
    """
    resolved_path = resolve_input_sheet(source)
    configs = _read_workbook(resolved_path)
    configs["_resolved_path"] = resolved_path
    configs["_is_temp"] = is_url(source)
    return configs
